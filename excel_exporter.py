# import required libraries
import os.path      # File I/O
from datetime import date
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill, Border, Side

import logger
import excel_new_file


EXCEL_OUTPUT_DIR = ''
EXCEL_DATA_ROW_START_INDEX = 1

# create new Excel file for export
def create_file(lines_of_data, filepath):
    try:
        # create a new file
        excel_new_file.create(filepath)
                
        return filepath
    
    except Exception, e:
        error_logger.log(e)

# exports to Excel file
def export(lines_of_data, filepath):
    try:
        if filepath is not None:
            # load workbook and get active sheet
            wb = load_workbook(filepath)
            sheet = wb.active

	    # row pointer
            row_index = EXCEL_DATA_ROW_START_INDEX

            border = Border(left=Side(style='hair', color='C8C8C8'),
                            right=Side(style='hair', color='C8C8C8'),
                            top=Side(style='hair', color='C8C8C8'),
                            bottom=Side(style='hair', color='C8C8C8'))

            # update Excel sheet row by row
            for data in lines_of_data:
                # add new data
                sheet.append(data)

                # colorize Excel cells
                for row_cells in sheet['A' + str(row_index):'D' + str(row_index)]:
                    for cell in row_cells:
                        cell.fill = PatternFill('solid', 'DFDEE8')
                        cell.border = border

                row_index += 1
            
            # save workbook
            wb.save(filepath)
            
    except Exception as e:
        logger.log_error(str(e))

