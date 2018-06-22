# import libraries
from openpyxl import load_workbook
from Tkinter import Tk
from tkFileDialog import askdirectory, askopenfilename
import os
import sys

# import logger module
import logger


# constants
EXCEL_DATA_ROW_START_INDEX = 1
CSV_OUTPUT_DIR = 'path/to/csv/dir/'

# let the user choose directory to save csv file
def choose_destination_dir():
    root = Tk()
    root.withdraw()
    root.update()
    directory = askdirectory()
    root.destroy()

    return directory


# let the user choose Excel file
def choose_excel_file():
    root = Tk()
    root.withdraw()
    root.update()
    filepath = askopenfilename()
    root.destroy()

    return filepath


# get rows from the file
def get_entries_from_file(filepath):
    entries = []
    wb = None

    try:
        # load workbook and get active sheet
        wb = load_workbook(filepath)
        sheet = wb.active

        # populate list with row entries
        for row in sheet.iter_rows(min_row=EXCEL_DATA_ROW_START_INDEX, max_row=sheet.max_row):
            # comma separated values from list
            client = ",".join(['"' + str(cell.value) + '"' for cell in row]) + "\n"
            entries.append(client)

    except IOError as e:
        error_logger.log(e)
        print 'Must provide Excel file to convert to .csv'
        entries = []

    except Exception as e:
        error_logger.log(e)
        print str(e)
        entries = []

    finally:
        if wb:
            wb.close()

    return entries


def create_csv(excel_file=None):
    try:
        if excel_file is None:
            # let the user choose csv dir
            excel_file = choose_excel_file()

            if not excel_file:
                print 'Must choose Excel file to convert!'
                return None

            directory = choose_destination_dir()

            if directory is None:
                print 'Aborting conversion process!'
                return None

            # path of csv file
            csv_file = directory + "/" + os.path.basename(excel_file).replace("xlsx", "csv")

        else:
            # make sure the directory exists
            if not os.path.exists(CSV_OUTPUT_DIR):
                try:
                    os.makedirs(CSV_OUTPUT_DIR)
                except:
                    pass

            # Destination directory is output/today's date/csv/
            # path of csv file
            csv_file = CSV_OUTPUT_DIR + os.path.basename(excel_file).replace("xlsx", "csv")

        # import data from Excel
        entries = get_entries_from_file(excel_file)

        # if record exists
        if len(entries) > 0:
            # write data to csv
            with open(csv_file, 'w') as f:
                f.writelines(entries)

    except Exception as e:
        print str(e)


