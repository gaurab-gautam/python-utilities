# import libraries
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

import logger	# see logger.py for its implementation


# constants
NAME_HEADER_WIDTH = 30
PHONE_HEADER_WIDTH = 15
EMAIL_HEADER_WIDTH = 30
ADDRESS_HEADER_WIDTH = 50
HEADER_ROW_HEIGHT = 10


# create headers and format cells
def create_sheet_header(sheet):
    # default font for the sheet
    default_font = Font(name='Times New Roman', bold=True, size=14)

    # merge cells
    # first line is empty
    sheet.merge_cells('A1:D1')
    # colorify
    sheet["A1"].fill = PatternFill('solid', '85929E')

    # set default font column width and column names
    # ...
    sheet.column_dimensions["A"].width = NAME_HEADER_WIDTH
    sheet["A2"].font = default_font
    sheet["A2"] = "Name"
    
    sheet.column_dimensions["B"].width = PHONE_HEADER_WIDTH
    sheet["B2"].font = default_font
    sheet["B2"] = "Phone Number"
    
    sheet.column_dimensions["C"].width = EMAIL_HEADER_WIDTH
    sheet["C2"].font = default_font
    sheet["C2"] = "Email"
    
    sheet.column_dimensions["D"].width = ADDRESS_HEADER_WIDTH
    sheet["D2"].font = default_font
    sheet["D2"] = "Address"
    
    
    # set row height for header row
    sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT
    sheet.row_dimensions[2].height = HEADER_ROW_HEIGHT
    
    return sheet
    

# create empty xl file at the given path 
def create(filepath):
    try:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Title"
        
        sheet = create_sheet_header(sheet)
        
        # save new Excel file
        wb.save(filepath)
        
    except Exception as e:
        logger.log_error(str(e))

    
