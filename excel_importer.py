# import module for Excel
import openpyxl


# imports from Excel file line by line
def import_from_excel(file_path):
    try:
        # load Excel workbook
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # data structure to hold list of search terms
        data = []
        
        # Get data line by line
        for i in range(1, sheet.max_row + 1):
            line = []
            for j in range(1, sheet.max_column + 1):
                cell_val = sheet.cell(row=i, column=j).value
                
                # ignore empty cells
                if cell_val is not None:
                    line.append(cell_val)
                
            data.append(line)    
            
    except Exception as e:
        data = None
        print(str(e))
        
    return data
            
            

